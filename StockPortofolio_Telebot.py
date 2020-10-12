import pandas as pd, telebot, datetime, time, copy
from openpyxl import load_workbook
from telebot import types

class User:
    def __init__(self, company_stock_symbol):
        self.company_stock_symbol = company_stock_symbol
        self.transaction = None
        self.stock_price = None
        self.shares_volume = None

API_TOKEN = "810975399:AAGWhxFmw_Cs96ZVZG3PeM9gTlz2nLTDpaM"
bot = telebot.TeleBot(API_TOKEN)

#Show available command(s) and bot information
@bot.message_handler(
    commands = [
        "help"
    ]
)
def helpService(message):
    msg = bot.reply_to(
        message, """This bot collects, counts, and stores your personal stock exchange activities.

Type /start to start using this bot and whenever you want to start from the beginning."""
    )
    
#Main feature
@bot.message_handler(
    commands = [
        "start"
    ]
)

def userStockCode(message):
    global todaysDate
    
    todaysDate = datetime.date.today().strftime("%d - %m - %Y").replace(" ", "")
    msg = bot.reply_to(
            message, "Please input YOUR USER STOCK CODE."
    )
    
    bot.register_next_step_handler(
            msg, initializeStockWorkbook
    )
        
def initializeStockWorkbook(message):
    global chat_id, stock_code, user
    
    chat_id = message.chat.id
    stock_code = message.text.upper()
    user = User(stock_code)

    try:
        load_workbook(
            r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
        )
    
    except:
        user_workbook = pd.ExcelWriter(
            r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
        )

        saldo = pd.DataFrame(
            columns = {
                "Date", "Bank Account", "Cash In", "Total"
                }
        )

        saldo.to_excel(
            user_workbook, sheet_name = "Saldo", index = False
        )
        
        user_workbook.save()
    
    markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
    )

    markup.add(
        "CASH TOP UP", "MARKET"
    )

    msg = bot.reply_to(
        message, "Please select your wanted transaction.", reply_markup = markup
    )
    
    bot.register_next_step_handler(
        msg, initializeTransactionType
    )

def initializeTransactionType(message):
    wanted_transaction = message.text

    if (wanted_transaction == "CASH TOP UP") or (wanted_transaction == "MARKET"):
        markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
        )
    
        markup.add(
            "TODAY'S DATE"
        )
    
        msg = bot.reply_to(
            message, """Please input the transaction date.
            
Select "TODAY'S DATE" if the transaction date is TODAY.
If NOT, input date manually  (dd-MM-yyyy).""", reply_markup = markup
        )
        
        if wanted_transaction == "CASH TOP UP":            
            bot.register_next_step_handler(
                msg, dateCashIn
            )
        
        else:
            bot.register_next_step_handler(
                msg, dateMarketTransaction
            )
    
    else:
        markup = types.ReplyKeyboardMarkup(
            one_time_keyboard = True
        )
    
        markup.add(
            "CASH TOP UP", "MARKET"
        )
        
        msg = bot.send_message(
            chat_id, "Please answer only from these below options.", reply_markup = markup
        )
    
        bot.register_next_step_handler(
            msg, initializeTransactionType
        )
        
        return

def dateCashIn(message):
    global topUpDate
    
    topUpDate = message.text.replace("/", "-").replace(".", "-")
    markup = types.ReplyKeyboardRemove()
    
    if topUpDate == "TODAY'S DATE":
        topUpDate = todaysDate
    
    else:
        if len(topUpDate) != 10:
            msg = bot.reply_to(
                message, "Please input the transaction date in the CORRECT FORMAT (dd-MM-yyyy).", reply_markup = markup
            )
        
            bot.register_next_step_handler(
                msg, dateCashIn
            )
            
            return
        
        else:
            pass
        
    msg = bot.reply_to(
        message, "Please input the bank's name.", reply_markup = markup
    )
    
    bot.register_next_step_handler(
        msg, bankAccount
    )

def bankAccount(message):
    global bank_account
    
    bank_account = message.text.upper().replace(",", "").replace(".", "")
    
    if not bank_account.isdigit():
        msg = bot.reply_to(
                message, "Please input the cash amount."
            )
        
        bot.register_next_step_handler(
            msg, topUpCash
        )
        
    else:
        msg = bot.reply_to(
                message, "Bank's name shouldn't contains number. Please input the name correctly."
            )
        
        bot.register_next_step_handler(
            msg, bankAccount
        )
        
        return

def topUpCash(message):
    global cash_amount
    
    cash_amount = message.text.replace(",", "").replace(".", "")
    
    if not cash_amount.isdigit():
        msg = bot.reply_to(
            message, "Cash input should be a number. Please re-input the amount."
        )
        
        bot.register_next_step_handler(
            msg, topUpCash
        )
        
        return
    
    markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
    )
    
    markup.add(
        "YES", "NO"
    )
    
    msg = bot.send_message(
                chat_id, f"""Are you sure this transaction is CORRECT?
                
Transaction date : {topUpDate}
Bank's Name : {bank_account}
Amount : {int(cash_amount):,}""", reply_markup = markup
    )
    
    bot.register_next_step_handler(
        msg, finalizeTopUpTransaction
    )

def finalizeTopUpTransaction(message):
    ensure = message.text
    
    if (ensure == "YES") or (ensure == "NO"):
        if ensure == "YES":
            user_workbook_xlsx = pd.ExcelFile(
                r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
            )
                
            balance = pd.read_excel(
                user_workbook_xlsx, "Saldo"
            )
            
            new_user_workbook = pd.ExcelWriter(
                r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx", engine = "openpyxl"
            )
            
            new_user_workbook.book = load_workbook(
                r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
            )
            
            new_user_workbook.sheets = dict(
                (
                    ws.title, ws
                ) for ws in new_user_workbook.book.worksheets
            )
    
            balance_new = pd.concat(
                [
                    balance, pd.DataFrame(
                        [
                            {
                                "Date" : topUpDate, "Bank Account" : bank_account, "Cash In" : f"{int(cash_amount):,}", "Total" : f"{balance['Cash In'].map(lambda x : int(x.replace(',', ''))).sum() + int(cash_amount):,}"
                            }
                        ]
                    )
                ]
            )
            
            balance_new["Date"] = pd.to_datetime(balance_new["Date"])
            
            balance_new.sort_values("Date", ascending = True).to_excel(
                        new_user_workbook, sheet_name = "Saldo", header = False, startrow = 1, index = False
                )
            
            new_user_workbook.save()
            
            bot.send_message(
                chat_id, """Saved to your workbook.

Type /start to begin another transaction."""
        )
            
        else:
            msg = bot.reply_to(
                message, "Let's start from the beginning. Please input YOUR USER STOCK CODE."
            )
            
            bot.register_next_step_handler(
                msg, initializeStockWorkbook
            )
                
            return
        
    else:
        markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
        )
        
        markup.add(
            "YES", "NO"
        )
        
        msg = bot.send_message(
            chat_id, "Please answer only from these below options.", reply_markup = markup
        )
                
        bot.register_next_step_handler(
            msg, finalizeTopUpTransaction
        )
                
        return

def dateMarketTransaction(message):
    global marketTransactionDate
    
    marketTransactionDate = message.text.replace("/", "-").replace(".", "-")
    markup = types.ReplyKeyboardRemove()
    
    if marketTransactionDate == "TODAY'S DATE":
        marketTransactionDate = todaysDate
    
    else:
        if len(marketTransactionDate) != 10:
            msg = bot.reply_to(
                message, "Please input the transaction date in the CORRECT FORMAT (dd-MM-yyyy).", reply_markup = markup
            )
        
            bot.register_next_step_handler(
                msg, dateMarketTransaction
            )
            
            return
        
        else:
            pass
        
    msg = bot.reply_to(
        message, "Please input the COMPANY STOCK SYMBOL.", reply_markup = markup
    )
    
    bot.register_next_step_handler(
        msg, initializeMarketTransaction
    )
    
def initializeMarketTransaction(message):
    company_stock_symbol = message.text.upper()

    user.company_stock_symbol = company_stock_symbol
    
    company_template_sheet = pd.DataFrame(
        columns = {
                "Date", "Transaction", "LOT", "Shares", "Stock Price", "Admin Fee", "Total", "Current LOT", "AVG"
        }
    )
    
    user_workbook = load_workbook(
        r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
    )
    
    if company_stock_symbol not in user_workbook.sheetnames:
        new_user_workbook = pd.ExcelWriter(
            r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx", engine = "openpyxl"
        )
        
        new_user_workbook.book = user_workbook
        
        new_user_workbook.sheets = dict(
                (
                    ws.title, ws
                ) for ws in new_user_workbook.book.worksheets
        )
        
        company_template_sheet.to_excel(
            new_user_workbook, sheet_name = company_stock_symbol, index = False
        )

        new_user_workbook.save()

    markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
    )

    markup.add(
        "BUY", "SELL"
    )

    msg = bot.reply_to(
        message, "Please select market transaction type.", reply_markup = markup
    )

    bot.register_next_step_handler(
        msg, marketTransactionType
    )

def marketTransactionType(message):
    transaction = message.text

    if (transaction == "BUY") or (transaction == "SELL"):
        user.transaction = transaction
            
    else:
        markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
        )
    
        markup.add(
            "BUY", "SELL"
        )
        
        msg = bot.send_message(
            chat_id, "Please answer only from these below options.", reply_markup = markup
        )

        bot.register_next_step_handler(
            msg, marketTransactionType
        )

        return

    msg = bot.reply_to(
        message, "Please input the stock price."
    )

    bot.register_next_step_handler(
        msg, stockPrice
    )

def stockPrice(message):
    stock_price = message.text.replace(".", "").replace(",", "")

    if not stock_price.isdigit():
        msg = bot.reply_to(
            message, "Price should be a number. Please input the stock price."
        )
        
        bot.register_next_step_handler(
            msg, stockPrice
        )
        
        return
    
    else:
        user.stock_price = stock_price
    
    msg = bot.reply_to(
        message, "Please input the amount of shares (LOT)."
    )
    
    bot.register_next_step_handler(
        msg, amountOfShares
    )

def amountOfShares(message):
    shares_volume = message.text.replace(".", "").replace(",", "")
    
    if not shares_volume.isdigit():
        msg = bot.reply_to(
                message, "Shares should be a number. Please input the amount of shares in LOT."
        )
        
        bot.register_next_step_handler(
                msg, amountOfShares
        )
        
        return
        
    else:
        user.shares_volume = shares_volume
    
    markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
    )
    
    markup.add(
        "YES", "NO"
    )
    
    msg = bot.send_message(
                chat_id, f"""Are you sure this transaction is CORRECT?

Transaction date : {marketTransactionDate}                
Stock : {user.company_stock_symbol}
Transaction Type : {user.transaction}
Price : {int(user.stock_price):,}
LOT : {int(user.shares_volume):,}""", reply_markup = markup
    )
    
    bot.register_next_step_handler(
        msg, finalizeMarketTransaction
    )

def finalizeMarketTransaction(message):
    ensure = message.text
        
    if (ensure == "YES") or (ensure == "NO"):
        if ensure == "YES":
            user_workbook_xlsx = pd.ExcelFile(
                    r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
            )
            
            css = pd.read_excel(
                    user_workbook_xlsx, user.company_stock_symbol
            )
            
            cssClone = copy.copy(css)
            
            cssClone["Total"] = cssClone["Total"].map(
                lambda x : float(
                    x.replace(",", "")
                    )
            )

            cssClone["Admin Fee"] = cssClone["Admin Fee"].map(
                lambda x : float(
                    x.replace(",", "")
                    )
            )
            
            currentLot = css[css.Transaction == "BUY"]["LOT"].sum() - css[css.Transaction == "SELL"]["LOT"].sum() + (-int(user.shares_volume) if user.transaction == "SELL" else int(user.shares_volume))
            
            adminFee = round(
                (
                    (
                        0.0017497439399112324 if user.transaction == "BUY" else 0.00275) * int(user.shares_volume) * 100 * int(user.stock_price)
                    ),
                2
            )
            
            avg = round(
                (
                    (
                        cssClone[cssClone.Transaction == "BUY"]["Total"].sum() - cssClone[cssClone.Transaction == "BUY"]["Admin Fee"].sum() + (
                            int(user.shares_volume) * 100 * int(user.stock_price)) 
                                                                                                                        )
                    /
                    (
                        css[css.Transaction == "BUY"]["Shares"].sum() + (int(user.shares_volume) * 100)
                        )
                ),
                2
                )            
            
            css_new = pd.concat(
                [
                    css, pd.DataFrame(
                    [
                        {
                    "Date" : marketTransactionDate, "Transaction" : user.transaction, "LOT" : format(int(user.shares_volume), ",d"), "Shares" : format(int(user.shares_volume) * 100, ",d"), "Stock Price" : format(int(user.stock_price), ",d"), "Admin Fee" : f"{adminFee:,}", "Total" : f"{((int(user.shares_volume) * 100 * int(user.stock_price)) + adminFee):,}",
                    "Current LOT" : format(currentLot, ",d"), "AVG" : f"{avg:,}" if user.transaction == "BUY" else ""
                            }
                        ]
                    )
                ]
            )
            
            css_new["Date"] = pd.to_datetime(css_new["Date"])
            
            new_user_workbook = pd.ExcelWriter(
                    r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx", engine = "openpyxl"
            )
            
            new_user_workbook.book = load_workbook(
                    r"C:\Users\Ezra\Desktop\\" + stock_code + ".xlsx"
            )
            
            new_user_workbook.sheets = dict(
                    (
                        ws.title, ws
                    ) for ws in new_user_workbook.book.worksheets
            )
            
            css_new.sort_values("Date", ascending = True).to_excel(
                    new_user_workbook, sheet_name = user.company_stock_symbol, header = False, startrow = 1, index = False
            )
            
            new_user_workbook.save()
            
            bot.send_message(
                chat_id, """Saved to your workbook.

Type /start to begin another transaction."""
            )
        
        else:
            msg = bot.reply_to(
                message, "Let's start from the beginning. Please input YOUR USER STOCK CODE."
            )
            
            bot.register_next_step_handler(
                msg, initializeStockWorkbook
            )
                
            return
        
    else:
        markup = types.ReplyKeyboardMarkup(
        one_time_keyboard = True
        )
        
        markup.add(
            "YES", "NO"
        )
        
        msg = bot.send_message(
            chat_id, "Please answer only from these below options.", reply_markup = markup
        )
                
        bot.register_next_step_handler(
            msg, finalizeMarketTransaction
        )
                
        return

while True:
    print (
        "Is running..."
    )
    
    try:
        bot.polling(
            none_stop = True
        )

    except Exception:
        time.sleep(5)